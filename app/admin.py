from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import login_user, logout_user, login_required, current_user, UserMixin
from app import login_manager, database, socketio
import firebase_admin
from firebase_admin import db
import pandas as pd
import io
import uuid
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

admin_bp = Blueprint('admin', __name__)

@admin_bp.before_request
def require_admin():
    if request.endpoint not in ['admin.login', 'admin.register'] and (not current_user.is_authenticated or not getattr(current_user, 'is_admin', False)):
        return redirect(url_for('admin.login'))

@admin_bp.route('/register', methods=['GET', 'POST'])
def register():
    # Check if any admin exists
    admin_users = database.child('users').get()
    admin_exists = False
    
    if admin_users:
        for uid, user in admin_users.items():
            if user.get('is_admin', False):
                admin_exists = True
                break
    
    if admin_exists:
        flash('Admin account already exists. Please login.', 'warning')
        return redirect(url_for('admin.login'))
    
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        fullname = request.form['fullname']
        
        # Check if email already exists
        existing_users = database.child('users').get()
        email_exists = False
        
        if existing_users:
            for uid, user in existing_users.items():
                if user.get('email') == email:
                    email_exists = True
                    break
        
        if email_exists:
            flash('Email already registered', 'danger')
            return render_template('admin/register.html')
        
        # Create new admin user
        uid = str(uuid.uuid4())
        new_user = {
            'id': uid,
            'email': email,
            'password': password,
            'fullname': fullname,
            'is_admin': True,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        # Add to database
        database.child('users').child(uid).set(new_user)
        
        flash('Admin account created successfully! Please login.', 'success')
        return redirect(url_for('admin.login'))
    
    return render_template('admin/register.html')

@admin_bp.route('/login', methods=['GET', 'POST'])
def login():
    # Check if any admin exists
    admin_users = database.child('users').get()
    admin_exists = False
    
    if admin_users:
        for uid, user in admin_users.items():
            if user.get('is_admin', False):
                admin_exists = True
                break
    
    if not admin_exists:
        flash('No admin account exists. Please create one.', 'warning')
        return redirect(url_for('admin.register'))
    
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        
        users = database.child('users').get()
        
        if users:
            for uid, user_data in users.items():
                if user_data.get('email') == email and user_data.get('is_admin', False) and user_data.get('password') == password:
                    user = UserMixin()
                    user.id = uid
                    user.email = email
                    user.is_admin = True
                    login_user(user)
                    flash('Admin login successful!', 'success')
                    return redirect(url_for('admin.dashboard'))
        flash('Invalid admin credentials', 'danger')
    return render_template('admin/login.html')

@admin_bp.route('/dashboard')
@login_required
def dashboard():
    # Get all users
    users_ref = database.child('users').get()
    users = users_ref if users_ref else {}
    
    # Get all matches
    matches_ref = database.child('matches').get()
    matches = matches_ref if matches_ref else {}
    
    # Get all reviews
    reviews_ref = database.child('reviews').get()
    reviews = reviews_ref if reviews_ref else {}
    
    # Get all career opportunities
    career_opportunities_ref = database.child('career_opportunities').get()
    career_opportunities = career_opportunities_ref if career_opportunities_ref else {}
    
    # Get all job applications
    job_applications_ref = database.child('job_applications').get()
    job_applications = job_applications_ref if job_applications_ref else {}
    
    # Get all chat messages
    chat_messages_ref = database.child('admin_chat').get()
    chat_messages = chat_messages_ref if chat_messages_ref else {}
    
    # Count unread messages for admin
    unread_count = 0
    user_messages = {}
    if chat_messages:
        for msg_id, msg in chat_messages.items():
            # Group messages by sender for better organization
            sender_id = msg.get('sender_id', '')
            if sender_id and sender_id != current_user.id:  # Not from admin
                if sender_id not in user_messages:
                    user_messages[sender_id] = []
                user_messages[sender_id].append({
                    'id': msg_id,
                    'message': msg.get('message', ''),
                    'timestamp': msg.get('timestamp', ''),
                    'sender_name': msg.get('sender_name', 'Unknown'),
                    'is_read': msg.get('is_read', False)
                })
                # Count unread messages
                if not msg.get('is_read', False):
                    unread_count += 1
    
    print(f"Admin dashboard: Found {len(chat_messages)} total messages, {unread_count} unread, from {len(user_messages)} users")
    
    # Get all unique job titles for filter options
    job_titles = set()
    education_levels = set()
    for uid, user in users.items():
        if user.get('is_admin', False):
            continue
        # Get job titles
        if user.get('job') and user.get('job').get('title'):
            job_titles.add(user.get('job').get('title').lower())
        # Get education levels
        if user.get('education') and user.get('education').get('level'):
            education_levels.add(user.get('education').get('level'))
    
    # Sort the job titles and education levels
    job_titles = sorted(job_titles)
    education_levels = sorted(education_levels)
    
    # Find potential matches based on gender preferences
    potential_matches = find_potential_matches(users, matches)
    
    # Ensure matches is a dictionary of dictionaries
    valid_matches = {}
    for mid, match in matches.items():
        if isinstance(match, dict) and 'user1_id' in match and 'user2_id' in match:
            valid_matches[mid] = match
    
    return render_template('admin/dashboard.html', users=users, matches=valid_matches, 
                           job_titles=job_titles, education_levels=education_levels,
                           potential_matches=potential_matches, reviews=reviews,
                           career_opportunities=career_opportunities, job_applications=job_applications,
                           chat_messages=chat_messages, user_messages=user_messages, unread_count=unread_count)

def find_potential_matches(users, existing_matches):
    """Find all potential male-female matches."""
    potential_matches = []
    
    # Extract existing match pairs to avoid duplicates
    existing_pairs = set()
    for match_id, match in existing_matches.items():
        if isinstance(match, dict) and 'user1_id' in match and 'user2_id' in match:
            user1_id = match.get('user1_id')
            user2_id = match.get('user2_id')
            existing_pairs.add((user1_id, user2_id))
            existing_pairs.add((user2_id, user1_id))  # Add reverse pair too
    
    # Group users by gender
    male_users = {}
    female_users = {}
    
    for uid, user in users.items():
        if user.get('is_admin', False):
            continue
            
        gender = user.get('gender', '').lower()
        if gender == 'male':
            male_users[uid] = user
        elif gender == 'female':
            female_users[uid] = user
    
    # Find all possible male-female combinations
    for male_id, male_user in male_users.items():
        for female_id, female_user in female_users.items():
            # Skip if this pair already exists in matches
            if (male_id, female_id) in existing_pairs:
                continue
                
            # Include all possible male-female combinations
            potential_matches.append({
                'male_id': male_id,
                'male_name': male_user.get('fullname', 'Unknown'),
                'female_id': female_id,
                'female_name': female_user.get('fullname', 'Unknown')
            })
    
    return potential_matches

# This route has been replaced by /careers/add route below

@admin_bp.route('/delete_career/<career_id>', methods=['POST'])
@login_required
def delete_career(career_id):
    """Delete a career opportunity."""
    try:
        database.child('career_opportunities').child(career_id).delete()
        flash('Career opportunity deleted successfully!', 'success')
    except Exception as e:
        flash(f'Error deleting career opportunity: {str(e)}', 'danger')

    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/send_admin_message', methods=['POST'])
@login_required
def send_admin_message():
    """Send a message from admin to user."""
    try:
        user_id = request.form['user_id']
        message = request.form['message']

        message_data = {
            'id': str(uuid.uuid4()),
            'sender_id': current_user.id,
            'receiver_id': user_id,
            'message': message,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'is_read': False,
            'type': 'admin_to_user'
        }

        database.child('admin_chat').push(message_data)
        
        # Emit Socket.IO event to notify user in real-time
        try:
            # Get admin name for the event
            admin_data = database.child('users').child(current_user.id).get()
            admin_name = admin_data.get('fullname', 'Admin') if admin_data else 'Admin'
            
            message_event_data = {
                'sender_id': current_user.id,
                'sender_name': admin_name,
                'receiver_id': user_id,
                'message': message,
                'timestamp': message_data['timestamp'],
                'is_read': False,
                'type': 'admin_to_user'
            }
            
            # Emit to user's room to update their chat
            socketio.emit('new_admin_message', message_event_data, room=user_id)
            print(f"Socket.IO event emitted to user {user_id} for message from admin {current_user.id}")
            
            # Also emit to admin's room so they can see their own message immediately
            socketio.emit('admin_message_sent', message_event_data, room=current_user.id)
            print(f"Socket.IO event emitted to admin {current_user.id} for their own sent message")
            
        except Exception as socket_error:
            print(f"Socket.IO error (non-critical): {str(socket_error)}")
            # Continue even if socket emission fails
        
        return jsonify({'status': 'success', 'message': 'Message sent successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@admin_bp.route('/download_xls')
@login_required
def download_xls():
    # Get users and matches data
    users_ref = database.child('users').get()
    users = users_ref if users_ref else {}
    
    matches_ref = database.child('matches').get()
    matches = matches_ref if matches_ref else {}
    
    # Get filter parameters
    job_filter = request.args.get('job', '').lower()
    education_filter = request.args.get('education', '').lower()
    location_filter = request.args.get('location', '').lower()
    
    # Prepare users data (excluding admins)
    users_data = []
    for uid, user in users.items():
        if user.get('is_admin', False):
            continue
        
        # Get education and job details
        education = user.get('education', {}) or {}
        job = user.get('job', {}) or {}
        preferences = user.get('preferences', {}) or {}
        
        # Apply filters if specified
        if job_filter and (not job.get('title') or job.get('title', '').lower() != job_filter):
            continue
            
        if education_filter and (not education.get('level') or education.get('level', '').lower() != education_filter):
            continue
            
        if location_filter:
            user_location = user.get('location', '').lower()
            if not user_location or location_filter not in user_location:
                continue
            
        user_data = {
            'Full Name': user.get('fullname', ''),
            'Gender': user.get('gender', ''),
            'Age': user.get('age', ''),
            'Location': user.get('location', ''),
            'Phone': user.get('phone', ''),
            'Profile Complete': 'Yes' if user.get('profile_complete', False) else 'No',
            # Education details
            'Education Level': education.get('level', ''),
            'Field of Study': education.get('field', ''),
            'Institution': education.get('institute', ''),
            # Job details
            'Job Title': job.get('title', ''),
            'Company': job.get('company', ''),
            'Experience (Years)': job.get('experience', ''),
            # Partner Preferences
            'Pref Age Min': preferences.get('age_min', ''),
            'Pref Age Max': preferences.get('age_max', ''),
            'Pref Location': preferences.get('location', ''),
            'Pref Education Level': preferences.get('education_level', ''),
            'Pref Education Field': preferences.get('education_field', ''),
            'Pref Job Title': preferences.get('job_title', ''),
            'Pref Min Experience': preferences.get('min_experience', ''),
            'Pref Qualities': preferences.get('qualities', '')
        }
        users_data.append(user_data)
    
    # Prepare matches data
    matches_data = []
    for match_id, match in matches.items():
        user1_id = match.get('user1_id', '')
        user2_id = match.get('user2_id', '')
        user1 = users.get(user1_id, {})
        user2 = users.get(user2_id, {})
        
        # Get job and education details for both users
        user1_job = user1.get('job', {}) or {}
        user2_job = user2.get('job', {}) or {}
        user1_edu = user1.get('education', {}) or {}
        user2_edu = user2.get('education', {}) or {}
        
        # Apply filters if specified for matches
        if job_filter:
            job1_match = user1_job.get('title', '').lower() and job_filter in user1_job.get('title', '').lower()
            job2_match = user2_job.get('title', '').lower() and job_filter in user2_job.get('title', '').lower()
            if not (job1_match or job2_match):
                continue
                
        if education_filter:
            edu1_match = user1_edu.get('level', '').lower() == education_filter
            edu2_match = user2_edu.get('level', '').lower() == education_filter
            if not (edu1_match or edu2_match):
                continue
        
        match_data = {
            'Male': user1.get('fullname', ''),
            'Male Phone': user1.get('phone', ''),
            'Male Job': user1_job.get('title', ''),
            'Male Company': user1_job.get('company', ''),
            'Male Education': user1_edu.get('level', ''),
            'Female': user2.get('fullname', ''),
            'Female Phone': user2.get('phone', ''),
            'Female Job': user2_job.get('title', ''),
            'Female Company': user2_job.get('company', ''),
            'Female Education': user2_edu.get('level', ''),
            'Match Score': match.get('match_score', ''),
            'Status': match.get('status', ''),
            'Match Date': match.get('match_date', '')
        }
        matches_data.append(match_data)
    
    # Create Excel file in memory
    output = io.BytesIO()
    
    # Create a Pandas Excel writer using XlsxWriter as the engine
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        # Convert users data to DataFrame and write to Excel
        if users_data:
            users_df = pd.DataFrame(users_data)
            users_df.to_excel(writer, sheet_name='Users', index=False)
        else:
            pd.DataFrame().to_excel(writer, sheet_name='Users', index=False)
        
        # Convert matches data to DataFrame and write to Excel
        if matches_data:
            matches_df = pd.DataFrame(matches_data)
            matches_df.to_excel(writer, sheet_name='Matches', index=False)
        else:
            pd.DataFrame().to_excel(writer, sheet_name='Matches', index=False)
    
    # Set up the response
    output.seek(0)
    
    # Generate filename with filters
    filename = 'user_data'
    if job_filter:
        filename += f'_job_{job_filter}'
    if education_filter:
        filename += f'_edu_{education_filter}'
    filename += '.xlsx'
    
    # Use send_file with mimetype and additional headers to ensure proper download behavior
    response = send_file(
        output, 
        download_name=filename, 
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Add a header to encourage redirect after download
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    response.headers["Location"] = url_for('admin.dashboard')
    
    # Flash a message that will be shown when they return to the dashboard
    flash('Data downloaded successfully!', 'success')
    
    return response

@admin_bp.route('/download_template')
@login_required
def download_template():
    """Download a sample Excel template for importing users."""
    try:
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "User Import Template"
        
        # Define the headers
        headers = [
            "Full Name", "Gender", "Age", "Location", "Phone", 
            "Education Level", "Field of Study", "Institution", 
            "Job Title", "Company", "Experience (Years)",
            "Pref Age Min", "Pref Age Max", "Pref Location", 
            "Pref Education Level", "Pref Education Field", 
            "Pref Job Title", "Min Experience", "Pref Qualities"
        ]
        
        # Define which columns are required
        required_cols = ["Full Name", "Gender"]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        required_fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
        optional_fill = PatternFill(start_color="6C757D", end_color="6C757D", fill_type="solid")
        
        # Add headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = required_fill if header in required_cols else optional_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Set column width
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
        
        # Add sample data
        sample_data = [
            ["John Doe", "Male", 30, "New York", "1234567890", 
             "Masters", "Computer Science", "MIT", 
             "Software Engineer", "Tech Corp", 5,
             25, 35, "New York", 
             "Bachelors", "Any", 
             "Any", 2, "Kind, Caring"],
            ["Jane Smith", "Female", 28, "Los Angeles", "0987654321", 
             "Bachelors", "Business", "UCLA", 
             "Marketing Manager", "Marketing Inc", 3,
             28, 38, "Any", 
             "Any", "Any", 
             "Any", 0, "Honest, Loyal"]
        ]
        
        # Add the sample data
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Add notes row
        notes_row = len(sample_data) + 3
        ws.cell(row=notes_row, column=1, value="Notes:").font = Font(bold=True)
        ws.cell(row=notes_row+1, column=1, value="1. Fields in blue are required")
        ws.cell(row=notes_row+2, column=1, value="2. Gender should be 'Male' or 'Female'")
        ws.cell(row=notes_row+3, column=1, value="3. Education Level options: High School, Bachelors, Masters, PhD")
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Set cache control headers to prevent caching
        response = send_file(
            output, 
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="user_import_template.xlsx"
        )
        
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        
        return response
        
    except Exception as e:
        print(f"Error generating template: {str(e)}")
        flash(f"Error generating template: {str(e)}", "danger")
        return redirect(url_for('admin.dashboard'))

@admin_bp.route('/create_match', methods=['POST'])
@login_required
def create_match():
    user1_id = request.form.get('user1_id')
    user2_id = request.form.get('user2_id')
    match_score = int(request.form.get('match_score', 0))
    
    if not user1_id or not user2_id:
        flash('Both users must be selected', 'danger')
        return redirect(url_for('admin.dashboard'))
    
    # Check if users exist
    users = database.child('users').get()
    
    if not users or user1_id not in users or user2_id not in users:
        flash('One or both users do not exist', 'danger')
        return redirect(url_for('admin.dashboard'))
    
    # Check if match already exists
    matches = database.child('matches').get() or {}
    
    for mid, match in matches.items():
        if (match.get('user1_id') == user1_id and match.get('user2_id') == user2_id) or \
           (match.get('user1_id') == user2_id and match.get('user2_id') == user1_id):
            flash('Match already exists between these users', 'warning')
            return redirect(url_for('admin.dashboard'))
    
    # Ensure user1 is Male and user2 is Female
    user1 = users.get(user1_id, {})
    user2 = users.get(user2_id, {})
    
    # If user1 is not male or user2 is not female, swap them
    if (user1.get('gender', '').lower() != 'male' or user2.get('gender', '').lower() != 'female'):
        if user1.get('gender', '').lower() == 'female' and user2.get('gender', '').lower() == 'male':
            # Swap user IDs
            user1_id, user2_id = user2_id, user1_id
        else:
            flash('Please select a male for the first user and a female for the second user', 'warning')
            return redirect(url_for('admin.dashboard'))
    
    # Create match
    match_id = str(uuid.uuid4())
    match_data = {
        'user1_id': user1_id,  # Male
        'user2_id': user2_id,  # Female
        'match_score': match_score,
        'status': 'pending',
        'match_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    
    database.child('matches').child(match_id).set(match_data)
    flash('Match created successfully', 'success')
    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/update_match_status/<match_id>', methods=['POST'])
@login_required
def update_match_status(match_id):
    new_status = request.form.get('status')
    
    if not new_status:
        flash('Status is required', 'danger')
        return redirect(url_for('admin.dashboard'))
    
    # Check if match exists
    match = database.child('matches').child(match_id).get()
    
    if not match:
        flash('Match not found', 'danger')
        return redirect(url_for('admin.dashboard'))
    
    # Update match status
    database.child('matches').child(match_id).update({'status': new_status})
    flash('Match status updated', 'success')
    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/delete_match/<match_id>', methods=['POST'])
@login_required
def delete_match(match_id):
    # Check if match exists
    match_ref = database.child('matches').child(match_id).get()
    
    if not match_ref:
        flash('Match not found.', 'danger')
        return redirect(url_for('admin.dashboard'))
    
    # Delete the match
    database.child('matches').child(match_id).delete()
    
    flash('Match deleted successfully.', 'success')
    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/create_all_matches', methods=['POST'])
@login_required
def create_all_matches():
    """Create all possible male-female matches."""
    # Get all users
    users_ref = database.child('users').get()
    users = users_ref if users_ref else {}
    
    # Get existing matches
    matches_ref = database.child('matches').get()
    matches = matches_ref if matches_ref else {}
    
    # Find all potential matches
    potential_matches = find_potential_matches(users, matches)
    
    # Create matches for all potential pairs
    matches_created = 0
    for match in potential_matches:
        male_id = match['male_id']
        female_id = match['female_id']
        
        # Create match
        match_id = str(uuid.uuid4())
        match_data = {
            'user1_id': male_id,  # Male
            'user2_id': female_id,  # Female
            'match_score': 50,  # Default score
            'status': 'pending',
            'match_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        database.child('matches').child(match_id).set(match_data)
        matches_created += 1
    
    if matches_created > 0:
        flash(f'Successfully created {matches_created} new matches!', 'success')
    else:
        flash('No new matches to create.', 'info')
    
    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/user/<user_id>', methods=['GET'])
@login_required
def view_user(user_id):
    """View detailed information about a specific user."""
    # Get user data
    user = database.child('users').child(user_id).get()
    
    if not user:
        flash('User not found', 'danger')
        return redirect(url_for('admin.dashboard'))
    
    # Get any matches this user is part of
    matches_ref = database.child('matches').get() or {}
    user_matches = []
    
    for match_id, match in matches_ref.items():
        if match.get('user1_id') == user_id or match.get('user2_id') == user_id:
            # Get the other user in this match
            other_user_id = match.get('user2_id') if match.get('user1_id') == user_id else match.get('user1_id')
            other_user = database.child('users').child(other_user_id).get() or {}
            
            user_matches.append({
                'match_id': match_id,
                'other_user': other_user,
                'other_user_id': other_user_id,
                'match_score': match.get('match_score', 0),
                'status': match.get('status', 'pending'),
                'match_date': match.get('match_date', '')
            })
    
    return render_template('admin/user_view.html', user=user, user_id=user_id, user_matches=user_matches)

@admin_bp.route('/import_users', methods=['POST'])
@login_required
def import_users():
    """Import users from Excel file and automatically match them."""
    if 'excel_file' not in request.files:
        flash('No file selected', 'danger')
        return redirect(url_for('admin.dashboard'))
        
    file = request.files['excel_file']
    
    if file.filename == '':
        flash('No file selected', 'danger')
        return redirect(url_for('admin.dashboard'))
        
    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('Invalid file format. Please upload an Excel file (.xlsx or .xls)', 'danger')
        return redirect(url_for('admin.dashboard'))
    
    try:
        # Read Excel file
        df = pd.read_excel(file)
        
        # Map Excel column names to database field names
        column_mapping = {
            'Full Name': 'fullname',
            'Gender': 'gender',
            'Age': 'age',
            'Location': 'location',
            'Phone': 'phone',
            'Education Level': 'education_level',
            'Field of Study': 'education_field',
            'Institution': 'education_institute',
            'Job Title': 'job_title',
            'Company': 'job_company',
            'Experience (Years)': 'job_experience',
            'Pref Age Min': 'pref_age_min',
            'Pref Age Max': 'pref_age_max',
            'Pref Location': 'pref_location',
            'Pref Education Level': 'pref_education_level',
            'Pref Education Field': 'pref_education_field',
            'Pref Job Title': 'pref_job_title',
            'Min Experience': 'pref_min_experience',
            'Pref Qualities': 'pref_qualities'
        }
        
        # Rename columns to match our expected format
        df = df.rename(columns=column_mapping)
        
        # Validate required columns
        required_columns = ['fullname', 'gender']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            flash(f'Missing required columns: {", ".join(missing_columns)}', 'danger')
            return redirect(url_for('admin.dashboard'))
        
        # Get existing users to avoid duplicates
        existing_users = database.child('users').get() or {}
        existing_emails = [user.get('email', '').lower() for uid, user in existing_users.items()]
        
        # Process each row
        users_added = 0
        users_skipped = 0
        
        for _, row in df.iterrows():
            # Skip rows with missing required fields
            if pd.isna(row['fullname']) or pd.isna(row['gender']):
                users_skipped += 1
                continue
                
            # Create new user
            uid = str(uuid.uuid4())
            
            # Prepare user data
            user_data = {
                'id': uid,
                'fullname': str(row['fullname']).strip(),
                'gender': str(row['gender']).lower().strip(),
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'is_admin': False,
                'profile_complete': False
            }
            
            # Add email if present
            if 'email' in df.columns and not pd.isna(row['email']):
                email = str(row['email']).lower().strip()
                
                # Skip if email already exists
                if email in existing_emails:
                    users_skipped += 1
                    continue
                
                user_data['email'] = email
                existing_emails.append(email)  # Add to list to prevent duplicates in same import
            else:
                # Generate a placeholder email if not provided
                user_data['email'] = f"user_{uid[:8]}@placeholder.com"
            
            # Add password if present
            if 'password' in df.columns and not pd.isna(row['password']):
                user_data['password'] = str(row['password']).strip()
            else:
                # Generate a random password if not provided
                user_data['password'] = str(uuid.uuid4())[:8]
            
            # Add optional fields if present
            optional_fields = ['age', 'location', 'phone']
            for field in optional_fields:
                if field in df.columns and not pd.isna(row[field]):
                    user_data[field] = str(row[field]).strip()
            
            # Add education if present
            education_fields = ['education_level', 'education_field', 'education_institute']
            if any(col in df.columns for col in education_fields):
                education_data = {}
                
                if 'education_level' in df.columns and not pd.isna(row['education_level']):
                    education_data['level'] = str(row['education_level']).strip()
                    
                if 'education_field' in df.columns and not pd.isna(row['education_field']):
                    education_data['field'] = str(row['education_field']).strip()
                    
                if 'education_institute' in df.columns and not pd.isna(row['education_institute']):
                    education_data['institute'] = str(row['education_institute']).strip()
                
                if education_data:
                    user_data['education'] = education_data
            
            # Add job if present
            job_fields = ['job_title', 'job_company', 'job_experience']
            if any(col in df.columns for col in job_fields):
                job_data = {}
                
                if 'job_title' in df.columns and not pd.isna(row['job_title']):
                    job_data['title'] = str(row['job_title']).strip()
                    
                if 'job_company' in df.columns and not pd.isna(row['job_company']):
                    job_data['company'] = str(row['job_company']).strip()
                    
                if 'job_experience' in df.columns and not pd.isna(row['job_experience']):
                    job_data['experience'] = str(row['job_experience']).strip()
                
                if job_data:
                    user_data['job'] = job_data
            
            # Add preferences if present
            preference_fields = ['pref_age_min', 'pref_age_max', 'pref_location', 'pref_education_level', 
                               'pref_education_field', 'pref_job_title', 'pref_min_experience', 'pref_qualities']
            
            if any(col in df.columns for col in preference_fields):
                preferences = {}
                
                field_mapping = {
                    'pref_age_min': 'age_min',
                    'pref_age_max': 'age_max',
                    'pref_location': 'location',
                    'pref_education_level': 'education_level',
                    'pref_education_field': 'education_field',
                    'pref_job_title': 'job_title',
                    'pref_min_experience': 'min_experience',
                    'pref_qualities': 'qualities'
                }
                
                for excel_field, db_field in field_mapping.items():
                    if excel_field in df.columns and not pd.isna(row[excel_field]):
                        preferences[db_field] = str(row[excel_field]).strip()
                
                if preferences:
                    user_data['preferences'] = preferences
            
            # Add user to database
            database.child('users').child(uid).set(user_data)
            users_added += 1
        
        # Create matches for new users
        if users_added > 0:
            # Get all users again including newly added ones
            all_users = database.child('users').get() or {}
            existing_matches = database.child('matches').get() or {}
            
            # Find potential matches
            potential_matches = find_potential_matches(all_users, existing_matches)
            
            # Create matches
            matches_created = 0
            for match in potential_matches:
                male_id = match['male_id']
                female_id = match['female_id']
                
                # Create match
                match_id = str(uuid.uuid4())
                match_data = {
                    'user1_id': male_id,  # Male
                    'user2_id': female_id,  # Female
                    'match_score': 50,  # Default score
                    'status': 'pending',
                    'match_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                
                database.child('matches').child(match_id).set(match_data)
                matches_created += 1
            
            flash(f'Successfully imported {users_added} users and created {matches_created} new matches! {users_skipped} users were skipped (duplicates or missing data).', 'success')
        else:
            flash(f'No new users were imported. {users_skipped} users were skipped (duplicates or missing data).', 'warning')
            
    except Exception as e:
        flash(f'Error importing users: {str(e)}', 'danger')
    
    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/delete_review/<review_id>', methods=['POST'])
@login_required
def delete_review(review_id):
    try:
        print(f"Attempting to delete review with ID: {review_id}")
        
        # Check if database is initialized
        if database is None:
            print("Error: Database is not initialized")
            flash('Database connection error.', 'danger')
            return redirect(url_for('admin.dashboard'))
            
        # Check if review exists
        review_ref = database.child('reviews').child(review_id).get()
        print(f"Review data: {review_ref}")
        
        if not review_ref:
            print(f"Review with ID {review_id} not found")
            flash('Review not found.', 'danger')
            return redirect(url_for('admin.dashboard'))
        
        # Delete the review - this is the critical operation
        try:
            print(f"Deleting review with ID: {review_id}")
            database.child('reviews').child(review_id).delete()
            print(f"Review with ID {review_id} deleted successfully from database")
            deletion_successful = True
        except Exception as db_error:
            print(f"Database error when deleting review: {str(db_error)}")
            flash('Error deleting review from database.', 'danger')
            return redirect(url_for('admin.dashboard'))
        
        # Emit socket event to update the index page in real-time
        # This is separated so that even if it fails, the review is still deleted
        try:
            print("Importing emit_review_deleted function")
            from app.socket_events import emit_review_deleted
            print("Emitting review_deleted event")
            emit_review_deleted(review_id)
            print("Event emitted successfully")
        except Exception as socket_error:
            print(f"Socket error (non-critical): {str(socket_error)}")
            # Continue even if socket emission fails
        
        # Return success message
        flash('Review deleted successfully.', 'success')
        return redirect(url_for('admin.dashboard'))
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        print(f"Error deleting review: {str(e)}")
        print(f"Traceback: {error_traceback}")
        flash('An error occurred while deleting the review.', 'danger')
        return redirect(url_for('admin.dashboard'))

@admin_bp.route('/test_reviews_db')
@login_required
def test_reviews_db():
    """Test route to check database connection for reviews."""
    try:
        # Check if database is initialized
        if database is None:
            return jsonify({
                'status': 'error',
                'message': 'Database is not initialized'
            }), 500
            
        # Try to get reviews
        reviews_ref = database.child('reviews').get()
        reviews = reviews_ref if reviews_ref else {}
        
        # Try to get a specific review
        if reviews:
            # Get the first review ID
            first_review_id = next(iter(reviews))
            first_review = database.child('reviews').child(first_review_id).get()
            
            # Try to remove and add back a test review
            test_review_id = 'test_review_id'
            test_review_data = {
                'user_name': 'Test User',
                'rating': 5,
                'comment': 'Test Comment',
                'date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            # Add test review
            database.child('reviews').child(test_review_id).set(test_review_data)
            
            # Get test review
            test_review = database.child('reviews').child(test_review_id).get()
            
            # Remove test review
            database.child('reviews').child(test_review_id).delete()
            
            return jsonify({
                'status': 'success',
                'message': 'Database connection test successful',
                'reviews_count': len(reviews),
                'first_review_id': first_review_id,
                'first_review': first_review,
                'test_review_added': test_review is not None
            })
        else:
            return jsonify({
                'status': 'warning',
                'message': 'No reviews found in database',
                'reviews_count': 0
            })
            
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        return jsonify({
            'status': 'error',
            'message': f'Error testing database: {str(e)}',
            'traceback': error_traceback
        }), 500

# Career management routes
@admin_bp.route('/careers/add', methods=['POST'])
@login_required
def add_career_opportunity():
    """Add a new career opportunity."""
    try:
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        requirements = request.form.get('requirements', '').strip()
        location = request.form.get('location', '').strip()
        job_type = request.form.get('job_type', 'full-time').strip()
        salary = request.form.get('salary', '').strip()
        deadline = request.form.get('deadline', '').strip()
        status = request.form.get('status', 'active').strip()
        
        if not title:
            flash('Job title is required', 'danger')
            return redirect(url_for('admin.dashboard'))
        
        # Create career opportunity
        career_id = str(uuid.uuid4())
        career_data = {
            'id': career_id,
            'title': title,
            'description': description,
            'requirements': requirements,
            'location': location,
            'job_type': job_type,
            'salary': salary,
            'deadline': deadline,
            'status': status,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'created_by': current_user.id
        }
        
        database.child('career_opportunities').child(career_id).set(career_data)
        flash('Career opportunity added successfully!', 'success')
        
    except Exception as e:
        flash(f'Error adding career opportunity: {str(e)}', 'danger')
    
    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/careers/edit/<career_id>', methods=['POST'])
@login_required
def edit_career(career_id):
    """Edit an existing career opportunity."""
    try:
        # Get existing career data
        career = database.child('career_opportunities').child(career_id).get()
        if not career:
            flash('Career opportunity not found', 'danger')
            return redirect(url_for('admin.dashboard'))
        
        # Update with form data
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        requirements = request.form.get('requirements', '').strip()
        location = request.form.get('location', '').strip()
        job_type = request.form.get('job_type', 'full-time').strip()
        salary = request.form.get('salary', '').strip()
        deadline = request.form.get('deadline', '').strip()
        status = request.form.get('status', 'active').strip()
        
        if not title:
            flash('Job title is required', 'danger')
            return redirect(url_for('admin.dashboard'))
        
        # Update career data
        career_data = {
            'id': career_id,
            'title': title,
            'description': description,
            'requirements': requirements,
            'location': location,
            'job_type': job_type,
            'salary': salary,
            'deadline': deadline,
            'status': status,
            'created_at': career.get('created_at', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            'created_by': career.get('created_by', current_user.id),
            'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'updated_by': current_user.id
        }
        
        database.child('career_opportunities').child(career_id).set(career_data)
        flash('Career opportunity updated successfully!', 'success')
        
    except Exception as e:
        flash(f'Error updating career opportunity: {str(e)}', 'danger')
    
    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/careers/delete/<career_id>', methods=['POST'])
@login_required
def delete_career_opportunity(career_id):
    """Delete a career opportunity."""
    try:
        # Check if career exists
        career = database.child('career_opportunities').child(career_id).get()
        if not career:
            flash('Career opportunity not found', 'danger')
            return redirect(url_for('admin.dashboard'))
        
        # Delete the career opportunity
        database.child('career_opportunities').child(career_id).delete()
        
        # Also delete all applications for this career
        applications_ref = database.child('job_applications').get()
        if applications_ref:
            for app_id, app_data in applications_ref.items():
                if app_data.get('career_id') == career_id:
                    database.child('job_applications').child(app_id).delete()
        
        flash('Career opportunity and all related applications deleted successfully!', 'success')
        
    except Exception as e:
        flash(f'Error deleting career opportunity: {str(e)}', 'danger')
    
    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/careers/<career_id>/applications')
@login_required
def view_career_applications(career_id):
    """View applications for a specific career opportunity."""
    try:
        # Get career data
        career = database.child('career_opportunities').child(career_id).get()
        if not career:
            flash('Career opportunity not found', 'danger')
            return redirect(url_for('admin.dashboard'))
        
        # Get all applications for this career
        applications_ref = database.child('job_applications').get()
        career_applications = {}
        
        if applications_ref:
            for app_id, app_data in applications_ref.items():
                if app_data.get('career_id') == career_id:
                    # Get user data for the applicant
                    user_id = app_data.get('user_id')
                    user_data = database.child('users').child(user_id).get() if user_id else {}
                    app_data['user_data'] = user_data
                    career_applications[app_id] = app_data
        
        return render_template('admin/career_applications.html', 
                               career=career, 
                               career_id=career_id,
                               applications=career_applications)
        
    except Exception as e:
        flash(f'Error viewing applications: {str(e)}', 'danger')
        return redirect(url_for('admin.dashboard'))

@admin_bp.route('/applications/<app_id>/status', methods=['POST'])
@login_required
def update_application_status(app_id):
    """Update the status of a job application."""
    try:
        new_status = request.form.get('status')
        if new_status not in ['approved', 'rejected', 'pending']:
            flash('Invalid status', 'danger')
            return redirect(request.referrer or url_for('admin.dashboard'))
        
        # Get the application
        application = database.child('job_applications').child(app_id).get()
        if not application:
            flash('Application not found', 'danger')
            return redirect(request.referrer or url_for('admin.dashboard'))
        
        # Update the status
        database.child('job_applications').child(app_id).update({
            'status': new_status,
            'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'updated_by': current_user.id
        })
        
        flash(f'Application status updated to {new_status}', 'success')
        
        # Redirect back to the applications page
        career_id = application.get('career_id')
        if career_id:
            return redirect(url_for('admin.view_career_applications', career_id=career_id))
        else:
            return redirect(url_for('admin.dashboard'))
        
    except Exception as e:
        flash(f'Error updating application status: {str(e)}', 'danger')
        return redirect(request.referrer or url_for('admin.dashboard'))
